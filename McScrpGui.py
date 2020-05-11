# -*- coding: utf-8 -*-
# 110 DESKTOP
# 159 SCRAPING
# 8 DJANGO
from threading import Thread
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QThread
from PyQt5.QtWidgets import QApplication, QWidget, QTableWidgetItem, QFileDialog, QStyleFactory, QDialog, QMessageBox, QPushButton
import sys
import requests
import McScrp
from openpyxl import Workbook


class Ui_Form(object):
    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(573, 430)
        icn = QtGui.QIcon('McScrp.ico')
        Form.setWindowIcon(icn)
        Form.setStyleSheet("""QPushButton {
  background-color: #4CAF50; /* Green */
  border: none;
  color: white;
  text-align: center;
  text-decoration: none;
  display: inline-block;
  font-size: 16px;
  -webkit-transition-duration: 0.4s; /* Safari */
  transition-duration: 0.4s;
  cursor: pointer;
}

QPushButton {
  background-color: white;
  color: black;
  border: 2px solid #555555;
}
QPushButton:pressed {
    background-color: #555555;
    color: white;
}
""")
        self.scrp = QtWidgets.QPushButton(Form)
        self.scrp.setGeometry(QtCore.QRect(640, 20, 110, 31))
        self.choice = QtWidgets.QComboBox(Form)
        self.choice.setGeometry(QtCore.QRect(530, 20, 101, 31))
        self.choice.addItems(["Text", "Tags"])
        self.save = QtWidgets.QPushButton(Form)
        self.save.setGeometry(QtCore.QRect(1200, 20, 121, 31))
        self.pbar = QtWidgets.QProgressBar(self)
        self.pbar.setGeometry(QtCore.QRect(790, 20, 350, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(50)
        self.scrp.setFont(font)
        self.choice.setFont(font)
        self.scrp.setObjectName("pushButton")
        font = QtGui.QFont()
        font.setPointSize(14)
        self.lineEdit_2 = QtWidgets.QLineEdit(Form)
        self.lineEdit_2.setGeometry(QtCore.QRect(20, 20, 500, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.lineEdit_2.setFont(font)
        self.lineEdit_2.setFocusPolicy(QtCore.Qt.ClickFocus)
        self.lineEdit_2.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.tbl = QtWidgets.QTableWidget(Form)
        self.tbl.setGeometry(QtCore.QRect(5, 71, 1355, 551))
        self.abt = QtWidgets.QPushButton(Form)
        self.abt.setGeometry(QtCore.QRect(640, 650, 91, 31))
        self.abt.setText('About Me')

        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "GuiMcScrp"))
        self.scrp.setText(_translate("Form", "Scraping data"))
        self.save.setText(_translate("Form", "Save data"))
        self.lineEdit_2.setPlaceholderText(
            _translate("Form", "URL Or Path Html"))
        QApplication.setStyle(QStyleFactory.create('Fusion'))

class MainWindow(QWidget, Ui_Form):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)
        QWidget.__init__(self)
        self.setupUi(self)
        self.scrp.clicked.connect(self.thr)
        self.save.clicked.connect(self.Save)
        self.abt.clicked.connect(self.abtme)

    def progress(self, datavalue):
        scr = McScrp.mcscrp()
        self.lis = scr.get_tags(datavalue)
        cc = self.lis.count("")
        if cc > 0:
            for i in range(cc):
                self.lis.remove("")

        self.tbl.setColumnCount(len(self.lis))
        self.tbl.setHorizontalHeaderLabels(self.lis)
        row = 1
        self.tbl.setRowCount(row)
        r = 1
        q = 0
        p = 0
        ch = {'Text': 'txt', 'Tags': 'tag'}
        for i in self.lis:
            resultat = scr.scrp(datavalue, i)[ch[self.choice.currentText()]]
            if len(resultat) > row:
                row = len(resultat)
                self.tbl.setRowCount(row)
            for u in resultat:
                if len(u) != 0:
                    self.tbl.setItem(r, q, QTableWidgetItem(str(u)))
                    r += 1
                    p += 1

                self.pbar.setValue(p)
            q += 1
            r = 0

    def abtme(self):
        QMessageBox(self).about(self, "About Me", '<h2 style="font-size: 20px; color: green;" >Contact me:</h2>  <a style="font-size:20px;" href="https://www.facebook.com/ZakatKnowledge/">Page Facebook</a>  <a style="font-size: 20px;" href="https://www.facebook.com/M97Chahboun">Account Facebook</a>  <a style="font-size: 20px; color: black;" href="https://www.github.com/ZakatKnowledge">Github</a>  <a style="font-size: 20px; color: deeppink;" href="https://www.instagram.com/zakat_of_knowledge/">Instagrame</a>  <a style="font-size: 20px; color: red;" href="https://youtube.com/channel/UCCiBkOPPs1iTCOyEeL7zWQg">Channel Youtube</a>        <h2 style="font-size: 20px; color: green;" >Developped By CHAHBOUN Mohammed</h2> ')


    def thr(self):
        Th = Thread(target=self.scraping)
        Th.start()
        Th.join()

    def scraping(self):
        self.tbl.clear()
        urlPath = self.lineEdit_2.text()
        if 'http' in urlPath:
            try:
                data = requests.get(urlPath).text
            except:
                self.error()
        elif 'html' in urlPath:
            try:
                data = open(urlPath, 'r').read()
            except FileNotFoundError:
                self.error()
        else:
            self.error()

        try:
            self.progress(data)
        except:
            pass

    def error(self):
        self.lineEdit_2.setText('')
        self.lineEdit_2.setPlaceholderText(
            "bad File Html or Url without http/s")

    def Save(self):
        file = QFileDialog.getSaveFileName(self, 'Enregistrer Fichier', 'data',
                                           ("Excel file (*.xlsx)"))
        wb = Workbook()
        ws = wb.active
        for u in range(len(self.lis)):
            ws.cell(column=u+1, row=1, value=self.lis[u])
        for w in range(self.tbl.rowCount()):
            for i in range(len(self.lis)):
                try:
                    ws.cell(column=i+1, row=w+2,
                            value=self.tbl.item(w, i).text())
                except:
                    pass
        wb.save(file[0])


def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.showMaximized()
    app.exec_()


if __name__ == '__main__':
    main()
